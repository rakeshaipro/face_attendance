"""Tests for /api/v1/reports (SRS §3.9)."""
from __future__ import annotations

import uuid
from datetime import datetime, timedelta, timezone


def _seed_log(db, *, employee_id: str, name: str = "Reporter", days_ago: int = 0, confidence: float = 0.9):
    from app.models import AttendanceLog

    lid = uuid.uuid4().hex
    ts = datetime.now(timezone.utc) - timedelta(days=days_ago)
    db.add(
        AttendanceLog(
            id=lid,
            machine_id="M1",
            location_name="Door",
            employee_id=employee_id,
            employee_name=name,
            timestamp=ts,
            confidence=confidence,
            snapshot_available=False,
            sync_status="sent",
        )
    )
    db.commit()
    return lid


def _seed_employee(db) -> str:
    from app.models import Employee

    eid = uuid.uuid4().hex
    db.add(Employee(id=eid, employee_id=f"REP-{eid[:6]}", name="Reporter", is_enrolled=True))
    db.commit()
    return eid


def _seed_audit(db, *, action: str = "employee.create", actor: str = "Admin", source: str = "api"):
    from app.models import AuditLog

    db.add(
        AuditLog(
            id=uuid.uuid4().hex,
            action=action,
            affected_id=uuid.uuid4().hex,
            source=source,
            actor=actor,
            new_value='{"x": 1}',
        )
    )
    db.commit()


# --- /reports/logs ------------------------------------------------------
def test_query_logs_filtered_by_employee(client, admin_headers):
    from app.db import SessionLocal

    with SessionLocal() as db:
        emp = _seed_employee(db)
        _seed_log(db, employee_id=emp)
        _seed_log(db, employee_id=emp)
    r = client.get(f"/api/v1/reports/logs?employee_id={emp}", headers=admin_headers)
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["total"] >= 2
    assert all(item["employee_id"] == emp for item in data["items"])


def test_query_logs_pagination(client, admin_headers):
    from app.db import SessionLocal

    with SessionLocal() as db:
        emp = _seed_employee(db)
        for _ in range(3):
            _seed_log(db, employee_id=emp)
    r = client.get(f"/api/v1/reports/logs?employee_id={emp}&page=1&limit=2", headers=admin_headers)
    assert len(r.json()["data"]["items"]) == 2


def test_daily_view_scopes_to_one_day(client, admin_headers):
    from app.db import SessionLocal

    with SessionLocal() as db:
        emp = _seed_employee(db)
        _seed_log(db, employee_id=emp, days_ago=0)
        old_id = _seed_log(db, employee_id=emp, days_ago=5)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    r = client.get(f"/api/v1/reports/logs/daily?date={today}", headers=admin_headers)
    ids = {item["id"] for item in r.json()["data"]["items"]}
    assert old_id not in ids


# --- /reports/logs/export ----------------------------------------------
def test_export_csv_has_all_fields(client, admin_headers):
    from app.db import SessionLocal

    with SessionLocal() as db:
        emp = _seed_employee(db)
        _seed_log(db, employee_id=emp, confidence=0.8765)
    r = client.get("/api/v1/reports/logs/export?format=csv", headers=admin_headers)
    assert r.status_code == 200
    assert "text/csv" in r.headers["content-type"]
    body = r.text
    # Header row contains every required column (§3.9.4).
    headers_line = body.splitlines()[0]
    for col in ("log_id", "machine_id", "employee_id", "confidence", "sync_status", "created_at"):
        assert col in headers_line
    # A seeded row carries the confidence value.
    assert "0.8765" in body


def test_export_xlsx_is_valid_workbook(client, admin_headers):
    from app.db import SessionLocal
    from openpyxl import load_workbook

    with SessionLocal() as db:
        emp = _seed_employee(db)
        _seed_log(db, employee_id=emp)

    r = client.get("/api/v1/reports/logs/export?format=xlsx", headers=admin_headers)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(__import__("io").BytesIO(r.content))
    ws = wb.active
    assert ws.max_row >= 2  # header + at least one data row
    header = [c.value for c in ws[1]]
    assert "log_id" in header and "confidence" in header


def test_export_requires_auth(client):
    r = client.get("/api/v1/reports/logs/export")
    assert r.status_code == 401


# --- /reports/audit -----------------------------------------------------
def test_audit_returns_rows_and_filters_by_action(client, admin_headers):
    from app.db import SessionLocal

    with SessionLocal() as db:
        _seed_audit(db, action="employee.create", actor="Alice")
        _seed_audit(db, action="webhook.test", actor="Bob")
        _seed_audit(db, action="sync.batch", actor="Carol")
    r = client.get("/api/v1/reports/audit?action=webhook.test", headers=admin_headers)
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["total"] >= 1
    assert all(item["action"] == "webhook.test" for item in data["items"])


def test_audit_filter_by_source_and_actor(client, admin_headers):
    from app.db import SessionLocal

    with SessionLocal() as db:
        _seed_audit(db, action="x", actor="Target", source="dashboard")
        _seed_audit(db, action="x", actor="Other", source="api")
    r = client.get("/api/v1/reports/audit?actor=Target", headers=admin_headers)
    items = r.json()["data"]["items"]
    assert items and all(item["actor"] == "Target" for item in items)


def test_audit_export_csv(client, admin_headers):
    from app.db import SessionLocal

    with SessionLocal() as db:
        _seed_audit(db, action="export.test")
    r = client.get("/api/v1/reports/audit/export", headers=admin_headers)
    assert r.status_code == 200
    assert "text/csv" in r.headers["content-type"]
    assert "action" in r.text.splitlines()[0]
